from openpyxl import load_workbook

def rowLen(sheet, column):
	for rowIndex in range(sheet.max_row, 0, -1):
		if sheet.cell(row = rowIndex, column = column).value != None:
			return rowIndex

def writeToJson(fileName, jsonEntries):
	with open(fileName, "w") as f:
		for jsonEntry in jsonEntries:
			f.write(jsonEntry)
			f.write("\n")

if __name__ == "__main__":

	wb = load_workbook(filename = 'spell_full.xlsx')

	columnNames = []

	sheet = wb["Updated 31Mar2020"]

	for columnIndex in range(sheet.max_column):
		columnNames.append(sheet.cell(row = 1, column = columnIndex + 1).value)

	jsonEntries = [""] * (sheet.max_row + 1)
	jsonEntries[0] = "["
	jsonEntries[len(jsonEntries) - 1] = "]"

	for j in range(1, sheet.max_row):
		jsonEntries[j] = jsonEntries[j] + "  {"
	
	for i, columnName in enumerate(columnNames):
		rowLength = rowLen(sheet, i + 1)

		for j in range(1, sheet.max_row):
			jsonEntries[j] = jsonEntries[j] + f" \"{sheet.cell(row = 1, column = i + 1).value}\": "

			if isinstance(sheet.cell(row = j + 1, column = i + 1).value, str):
				value = sheet.cell(row = j + 1, column = i + 1).value.replace("\"","\\\"")
				jsonEntries[j] = jsonEntries[j] + f"\"{value}\","
			
			elif sheet.cell(row = j + 1, column = i + 1).value is None:
				jsonEntries[j] = jsonEntries[j] + "null,"
			
			else:
				jsonEntries[j] = jsonEntries[j] + f"{sheet.cell(row = j + 1, column = i + 1).value},"

	for j in range(1, sheet.max_row):
		jsonEntries[j] = jsonEntries[j][:-1]	#gets rid of last , for the key pair values
		jsonEntries[j] = jsonEntries[j] + " },"

	jsonEntries[len(jsonEntries) - 2] = jsonEntries[len(jsonEntries) - 2][:-1]	#gets rid of last , for the }

	writeToJson("spells.json", jsonEntries)	
